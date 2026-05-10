import html
from datetime import datetime
from io import BytesIO
from pathlib import Path
from typing import Any
from urllib.parse import quote

from pydantic import BaseModel
from fastapi import APIRouter, HTTPException, Query, Request, Response


class DifyAnalyzeRequest(BaseModel):
    file_name: str
    capture_time: str | None = None
    zone_id: str | None = None
    yolo_result: str | None = None
    confidence: int | None = None
    description: str | None = None


def _severity_text(value: Any) -> str:
    if value == "danger":
        return "严重"
    if value == "warn":
        return "异常"
    if value in {"good", "healthy"}:
        return "健康"
    return "健康"


def _confidence_percent(value: Any) -> str:
    try:
        return f"{float(value) * 100:.1f}%"
    except (TypeError, ValueError):
        return "--"


def _format_detections(record: dict[str, Any]) -> str:
    detections = record.get("detections") or []
    if not detections:
        return "无"
    parts = []
    for item in detections:
        label = item.get("label") or "--"
        conf = _confidence_percent(item.get("confidence"))
        parts.append(f"{label}({conf})")
    return "；".join(parts)


def _pdf_text(value: Any) -> str:
    return html.escape(str(value), quote=False).replace("\n", "<br/>")


def _build_export_rows(records: list[dict[str, Any]]) -> list[list[str]]:
    rows = [
        [
            "时间",
            "区域",
            "目标数量",
            "对应病症",
        ]
    ]
    for record in records:
        rows.append(
            [
                str(record.get("capture_time") or "--"),
                str(record.get("zone_id") or "--"),
                str(record.get("detection_count") or 0),
                _format_detections(record),
            ]
        )
    return rows


def _build_xlsx(records: list[dict[str, Any]], output_dir: Path, annotated_dir: Path) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    generated_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    abnormal_count = sum(1 for r in records if r.get("severity") != "good")

    wb = Workbook()
    ws = wb.active
    ws.title = "YOLO导出"

    title_fill = PatternFill("solid", fgColor="1F6F4A")
    header_fill = PatternFill("solid", fgColor="DDEFE6")
    soft_fill = PatternFill("solid", fgColor="F5FAF7")
    white_font = Font(name="Microsoft YaHei", color="FFFFFF", bold=True, size=16)
    normal_font = Font(name="Microsoft YaHei", size=11, color="26352D")
    header_font = Font(name="Microsoft YaHei", bold=True, size=11, color="1F3B2D")
    thin = Side(style="thin", color="C7D8CE")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.merge_cells("A1:D1")
    ws["A1"] = "红橙果园 YOLO 视觉诊断导出"
    ws["A1"].fill = title_fill
    ws["A1"].font = white_font
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    summary_rows = [
        ("生成时间", generated_at),
        ("图片数量", f"{len(records)} 张"),
        ("异常/预警数量", f"{abnormal_count} 处"),
    ]
    for row_idx, (label, value) in enumerate(summary_rows, start=2):
        ws.cell(row_idx, 1, label)
        ws.cell(row_idx, 2, value)
        ws.merge_cells(start_row=row_idx, start_column=2, end_row=row_idx, end_column=4)
        for col in range(1, 5):
            cell = ws.cell(row_idx, col)
            cell.font = header_font if col == 1 else normal_font
            cell.fill = soft_fill
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
        ws.row_dimensions[row_idx].height = 24

    rows = _build_export_rows(records)
    detail_start_row = 6
    for row_offset, row in enumerate(rows):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(detail_start_row + row_offset, col_idx, value)

    ws.freeze_panes = f"A{detail_start_row + 1}"
    ws.auto_filter.ref = f"A{detail_start_row}:D{detail_start_row + max(0, len(rows) - 1)}"
    widths = [22, 14, 12, 72]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    for row_idx in range(detail_start_row, detail_start_row + len(rows)):
        row = ws[row_idx]
        for cell in row:
            cell.font = header_font if row_idx == detail_start_row else normal_font
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if row_idx == detail_start_row:
                cell.fill = header_fill
        ws.row_dimensions[row_idx].height = 26 if row_idx == detail_start_row else 58

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _record_image_path(record: dict[str, Any], output_dir: Path, annotated_dir: Path) -> Path | None:
    annotated_name = record.get("annotated_file_name")
    if annotated_name:
        annotated_path = annotated_dir / str(annotated_name)
        if annotated_path.exists():
            return annotated_path
    file_name = record.get("file_name")
    if file_name:
        image_path = output_dir / str(file_name)
        if image_path.exists():
            return image_path
    return None


def _build_pdf(records: list[dict[str, Any]], output_dir: Path, annotated_dir: Path) -> bytes:
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.cidfonts import UnicodeCIDFont
    from reportlab.platypus import Image as PdfImage
    from reportlab.platypus import KeepTogether, PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    pdfmetrics.registerFont(UnicodeCIDFont("STSong-Light"))

    generated_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    abnormal_count = sum(1 for r in records if r.get("severity") != "good")
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=14 * mm,
        leftMargin=14 * mm,
        topMargin=14 * mm,
        bottomMargin=14 * mm,
        title="红橙果园 YOLO 视觉诊断植保报告",
    )

    base = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "CnTitle",
        parent=base["Title"],
        fontName="STSong-Light",
        fontSize=20,
        leading=28,
        textColor=colors.HexColor("#1F6F4A"),
        alignment=TA_CENTER,
        spaceAfter=8,
    )
    sub_style = ParagraphStyle(
        "CnSub",
        parent=base["Normal"],
        fontName="STSong-Light",
        fontSize=10,
        leading=15,
        textColor=colors.HexColor("#5B6B62"),
        alignment=TA_CENTER,
        spaceAfter=12,
    )
    normal_style = ParagraphStyle(
        "CnNormal",
        parent=base["Normal"],
        fontName="STSong-Light",
        fontSize=9,
        leading=13,
        textColor=colors.HexColor("#26352D"),
    )
    section_style = ParagraphStyle(
        "CnSection",
        parent=normal_style,
        fontSize=13,
        leading=18,
        textColor=colors.HexColor("#1F6F4A"),
        spaceBefore=10,
        spaceAfter=6,
    )

    story: list[Any] = [
        Paragraph("红橙果园 YOLO 视觉诊断植保报告", title_style),
        Paragraph(_pdf_text(f"生成时间：{generated_at}"), sub_style),
    ]

    summary_data = [
        ["图片数量", "异常/预警数量", "健康样本", "生成时间"],
        [f"{len(records)} 张", f"{abnormal_count} 处", f"{len(records) - abnormal_count} 张", generated_at],
    ]
    summary_table = Table(summary_data, colWidths=[34 * mm, 34 * mm, 34 * mm, 64 * mm])
    summary_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F6F4A")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("BACKGROUND", (0, 1), (-1, 1), colors.HexColor("#F2FAF5")),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#BFD5C7")),
                ("FONTNAME", (0, 0), (-1, -1), "STSong-Light"),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.HexColor("#F2FAF5")]),
            ]
        )
    )
    story.extend([summary_table, Spacer(1, 8 * mm), Paragraph("诊断明细", section_style)])

    if not records:
        story.append(Paragraph("当前 output 文件夹下暂无可导出的模拟预测数据。", normal_style))
    for idx, record in enumerate(records, start=1):
        image_path = _record_image_path(record, output_dir, annotated_dir)
        thumb: Any = Paragraph("无图片", normal_style)
        if image_path:
            try:
                thumb = PdfImage(str(image_path), width=45 * mm, height=35 * mm, kind="proportional")
            except Exception:
                thumb = Paragraph("图片读取失败", normal_style)

        info = [
            [Paragraph("图片文件", normal_style), Paragraph(_pdf_text(record.get("file_name") or "--"), normal_style)],
            [Paragraph("识别时间", normal_style), Paragraph(_pdf_text(record.get("capture_time") or "--"), normal_style)],
            [Paragraph("区域", normal_style), Paragraph(_pdf_text(record.get("zone_id") or "--"), normal_style)],
            [
                Paragraph("诊断结果", normal_style),
                Paragraph(
                    _pdf_text(
                        f"{record.get('summary_label') or '--'} / {_confidence_percent(record.get('summary_confidence'))} / {_severity_text(record.get('severity'))}"
                    ),
                    normal_style,
                ),
            ],
            [Paragraph("目标数量", normal_style), Paragraph(str(record.get("detection_count") or 0), normal_style)],
            [Paragraph("说明", normal_style), Paragraph(_pdf_text(record.get("description") or "--"), normal_style)],
            [Paragraph("检出目标", normal_style), Paragraph(_pdf_text(_format_detections(record)), normal_style)],
        ]
        detail_table = Table(info, colWidths=[20 * mm, 93 * mm])
        detail_table.setStyle(
            TableStyle(
                [
                    ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#D2E2D8")),
                    ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#EEF7F2")),
                    ("FONTNAME", (0, 0), (-1, -1), "STSong-Light"),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("LEFTPADDING", (0, 0), (-1, -1), 5),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 5),
                    ("TOPPADDING", (0, 0), (-1, -1), 4),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                ]
            )
        )
        card = Table(
            [[Paragraph(_pdf_text(f"{idx}. {record.get('file_name') or '--'}"), section_style)], [Table([[thumb, detail_table]], colWidths=[49 * mm, 113 * mm])]],
            colWidths=[166 * mm],
        )
        card.setStyle(
            TableStyle(
                [
                    ("BOX", (0, 0), (-1, -1), 0.6, colors.HexColor("#AFCDB9")),
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#E5F3EA")),
                    ("LEFTPADDING", (0, 0), (-1, -1), 6),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 6),
                    ("TOPPADDING", (0, 0), (-1, -1), 5),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
                ]
            )
        )
        story.append(KeepTogether([card, Spacer(1, 5 * mm)]))
        if idx % 4 == 0 and idx < len(records):
            story.append(PageBreak())

    doc.build(story)
    return buffer.getvalue()


def build_api_router(
    data_service: Any,
    yolo_service: Any,
    dify_service: Any,
    output_dir: Path,
    valid_zones: list[str],
    metric_keys: list[str],
) -> APIRouter:
    router = APIRouter(prefix="/api")

    def normalize_zone(value: str) -> str:
        return value.strip().lower().replace("-", "_")

    def ensure_login(request: Request) -> None:
        if not request.session.get("user"):
            raise HTTPException(status_code=401, detail="not authenticated")

    @router.get("/zones")
    def api_zones(request: Request) -> dict[str, Any]:
        ensure_login(request)
        return {"zones": data_service.build_zone_cards()}

    @router.get("/latest")
    def api_latest(request: Request, zone_id: str = Query("zone_1")) -> dict[str, Any]:
        ensure_login(request)
        zone = normalize_zone(zone_id)
        if zone not in valid_zones:
            raise HTTPException(status_code=400, detail="invalid zone_id")

        row = data_service.latest_zone_row(zone)
        if row is None:
            raise HTTPException(status_code=404, detail="no sensor data")

        row["timestamp"] = row["timestamp"].strftime("%Y-%m-%d %H:%M:%S")
        return {
            "zone_id": zone,
            "timestamp": row["timestamp"],
            "metrics": {key: row.get(key) for key in metric_keys},
            "raw": row,
        }

    @router.get("/predictions")
    def api_predictions(
        request: Request,
        zone_id: str = Query("zone_1"),
        limit: int = Query(24, ge=1, le=72),
    ) -> dict[str, Any]:
        ensure_login(request)
        zone = normalize_zone(zone_id)
        if zone not in valid_zones:
            raise HTTPException(status_code=400, detail="invalid zone_id")

        rows = data_service.latest_predictions(zone, limit=limit)
        for row in rows:
            row["predict_time"] = row["predict_time"].strftime("%Y-%m-%d %H:%M:%S")

        return {"zone_id": zone, "count": len(rows), "rows": rows}

    @router.get("/yolo-detections")
    def api_yolo_detections(
        request: Request,
        limit: int = Query(5, ge=1, le=200),
        auto_refresh: bool = Query(False),
    ) -> dict[str, Any]:
        ensure_login(request)
        # For big-screen simulation, optionally rescan output images before reading latest rows.
        if auto_refresh:
            yolo_service.refresh_all()
        records = yolo_service.latest_detections(limit=limit)
        return {"records": records, "count": len(records)}

    @router.get("/yolo-export/excel")
    def api_yolo_export_excel(request: Request) -> Response:
        ensure_login(request)
        records = yolo_service.refresh_all()
        content = _build_xlsx(records, output_dir, yolo_service.annotated_dir)
        filename = f"yolo_detections_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return Response(
            content=content,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}; filename*=UTF-8''{quote(filename)}"},
        )

    @router.get("/yolo-export/pdf")
    def api_yolo_export_pdf(request: Request) -> Response:
        ensure_login(request)
        records = yolo_service.refresh_all()
        content = _build_pdf(records, output_dir, yolo_service.annotated_dir)
        filename = f"yolo_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        return Response(
            content=content,
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename={filename}; filename*=UTF-8''{quote(filename)}"},
        )

    @router.get("/yolo-annotated")
    def api_yolo_annotated(request: Request, file_name: str = Query(...)) -> dict[str, Any]:
        ensure_login(request)
        try:
            annotated = yolo_service.build_annotated_image(file_name)
        except FileNotFoundError:
            raise HTTPException(status_code=404, detail="annotated image source not found")
        return {"annotated_image_url": f"/annotated/{annotated.name}"}

    @router.post("/yolo-refresh")
    def api_yolo_refresh(request: Request, file_name: str | None = Query(default=None)) -> dict[str, Any]:
        ensure_login(request)
        if file_name:
            try:
                record = yolo_service.refresh_one(file_name)
            except FileNotFoundError:
                raise HTTPException(status_code=404, detail="image not found in output directory")
            return {"mode": "single", "record": record}

        records = yolo_service.refresh_all()
        return {"mode": "all", "count": len(records)}

    @router.post("/dify/analyze-image")
    def api_dify_analyze_image(request: Request, body: DifyAnalyzeRequest) -> dict[str, Any]:
        ensure_login(request)
        if not dify_service.enabled:
            raise HTTPException(status_code=503, detail="dify api key not configured")

        image_path = output_dir / body.file_name
        if not image_path.exists() or not image_path.is_file():
            raise HTTPException(status_code=404, detail="image not found")

        try:
            annotated_image_path = yolo_service.build_annotated_image(body.file_name)
        except FileNotFoundError as exc:
            raise HTTPException(status_code=404, detail=str(exc))

        user = request.session.get("user") or "dashboard-user"
        prompt = (
            "你是一名植保专家，请根据这张YOLO带框图快速给出结论。"
            f"\n时间: {body.capture_time or '--'}"
            f"\n区域: {body.zone_id or '--'}"
            f"\nYOLO结果: {body.yolo_result or '--'}"
            f"\n置信度: {body.confidence if body.confidence is not None else '--'}%"
            f"\n补充说明: {body.description or '--'}"
            "\n请用中文直接输出，不要输出思考过程。"
            "\n格式：1) 病虫害判断 2) 严重程度 3) 用药/处置建议 4) 复查建议。"
        )
        try:
            result = dify_service.analyze_image(image_path=annotated_image_path, prompt=prompt, user=str(user))
        except FileNotFoundError:
            raise HTTPException(status_code=404, detail="image not found")
        except Exception as exc:
            raise HTTPException(status_code=502, detail=f"dify request failed: {exc}")

        return result

    @router.get("/yolo-placeholder")
    def api_yolo_placeholder(request: Request, limit: int = Query(4, ge=1, le=12)) -> dict[str, Any]:
        ensure_login(request)
        records = yolo_service.latest_detections(limit=limit)
        return {"records": records}

    @router.get("/overview")
    def api_overview(request: Request, zone_id: str = Query("zone_1")) -> dict[str, Any]:
        ensure_login(request)
        zone = normalize_zone(zone_id)
        if zone not in valid_zones:
            raise HTTPException(status_code=400, detail="invalid zone_id")

        latest_row = data_service.latest_zone_row(zone)
        latest_payload = None
        if latest_row:
            latest_payload = {
                "timestamp": latest_row["timestamp"].strftime("%Y-%m-%d %H:%M:%S"),
                "metrics": {key: latest_row.get(key) for key in metric_keys},
            }

        prediction_rows = data_service.latest_predictions(zone, limit=24)
        for row in prediction_rows:
            row["predict_time"] = row["predict_time"].strftime("%Y-%m-%d %H:%M:%S")

        return {
            "now": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "zone_id": zone,
            "zones": data_service.build_zone_cards(),
            "latest": latest_payload,
            "predictions": prediction_rows,
            "yolo_records": yolo_service.latest_detections(limit=5),
        }

    return router
